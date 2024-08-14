from openpyxl import load_workbook
import os

file_name = "E://tu_hoc//excel//test_canh_bao_hoc_vu.xlsx"
hk_hien_tai = 3
nam_hoc_hien_tai = 2023
nam_hoc_hk_hien_tai = str(nam_hoc_hien_tai) + str(hk_hien_tai)

wb_obj = load_workbook(file_name)
sheet_obj = wb_obj.active

number_of_row = sheet_obj.max_row
number_of_col = sheet_obj.max_column

for r in range(2, number_of_row + 1):
    nam_hoc_hk_vao = str(sheet_obj.cell(r, 1).value)
    so_hk_toi_da = sheet_obj.cell(r, 2).value
    nam_hoc = int(nam_hoc_hk_vao[0:4])
    hoc_ky = int(nam_hoc_hk_vao[-1])
    for i in range(1, so_hk_toi_da + 1):
        if hoc_ky == 3:
            hoc_ky = 1
            nam_hoc += 1
        else:
            hoc_ky += 1
    nam_hoc_hoc_ky_het_han = str(nam_hoc) + str(hoc_ky)
    sheet_obj.cell(r, 3).value = nam_hoc_hoc_ky_het_han
    hoc_ky_con_lai = 0
    while(1):
        if hk_hien_tai == 3:
            hk_hien_tai = 1
            nam_hoc_hien_tai += 1
        else:
            hk_hien_tai += 1

        nam_hoc_hk_con_lai = str(nam_hoc_hien_tai) + str(hk_hien_tai)
        if nam_hoc_hk_con_lai != nam_hoc_hoc_ky_het_han:
            hoc_ky_con_lai += 1
        else:
            break
    print(f"năm học học kỳ hết hạn {nam_hoc_hoc_ky_het_han}")
    print(f"năm học học kỳ hiện tại {nam_hoc_hk_hien_tai}")
    print(f"Số học kỳ còn lại {hoc_ky_con_lai}")
    sheet_obj.cell(r, 4).value = hoc_ky_con_lai
    #print(f"năm học hết hạn: {nam_hoc} - {nam_hoc + 1} học kỳ cuối: {hoc_ky}") 
    break
#wb_obj.save(file_name) 